from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

PRIMARY_SHEETS = ("完整策略表", "补充策略")
DETAIL_SHEETS = ("Sheet1", "Sheet2")

DETAIL_ALIASES = {
    "买入看涨": "买入看涨",
    "买入看跌": "买入看跌",
    "卖出备兑看涨": "备兑看涨",
    "卖出裸看涨": "卖出看涨",
    "卖出裸看跌": "卖出看跌",
    "卖出备兑看跌": "备兑看跌",
    "牛市看涨价差": "牛市看涨价差",
    "牛市看跌价差": "牛市看跌价差",
    "牛市看涨跨期价差": "看涨日历价差",
    "熊市看跌价差": "熊市看跌价差",
    "熊市看涨价差": "熊市看涨价差",
    "熊市看跌跨期价差": "看跌日历价差",
    "买入跨式价差": "买入跨式",
    "卖出跨式价差": "卖出跨式",
    "买入宽跨式价差": "买入宽跨式",
    "卖出宽跨式价差": "卖出宽跨式",
    "铁鹰式价差": "铁鹰式",
    "看涨期权反向价差": "看涨反向比率价差",
    "领圈": "领口策略",
    "盒式价差套利": "盒式套利",
    "反向转换套利": "反转套利",
}


def normalize_strategy_name(name: str) -> str:
    cleaned = re.sub(r"[（(].*?[）)]", "", name)
    cleaned = cleaned.replace("期权", "")
    cleaned = cleaned.replace("策略", "")
    cleaned = cleaned.replace(" ", "")
    cleaned = cleaned.replace("　", "")
    cleaned = cleaned.strip()
    return cleaned


def split_strategy_name(name: str) -> tuple[str, str]:
    match = re.match(r"^(.*?)(?:[（(](.*?)[）)])?$", name.strip())
    if not match:
        return name.strip(), ""
    chinese_name = match.group(1).strip()
    english_name = (match.group(2) or "").strip()
    return chinese_name, english_name


def dedupe(values: list[str]) -> list[str]:
    seen: set[str] = set()
    ordered: list[str] = []
    for value in values:
        if not value or value in seen:
            continue
        ordered.append(value)
        seen.add(value)
    return ordered


def risk_level_from_label(label: str) -> int:
    match = re.search(r"(\d+)\s*级", label)
    return int(match.group(1)) if match else 3


def risk_bucket(level: int) -> str:
    if level <= 2:
        return "低风险"
    if level == 3:
        return "中风险"
    return "高风险"


def strategy_family(name: str) -> str:
    if "套利" in name:
        return "套利"
    if any(keyword in name for keyword in ("保护性", "备兑", "领口", "双限")):
        return "对冲/收益增强"
    if "合成" in name:
        return "合成复制"
    if any(keyword in name for keyword in ("对角", "日历", "跨期")):
        return "时间价差"
    if any(keyword in name for keyword in ("跨式", "宽跨式", "蝶", "鹰", "Strap", "Strip")):
        return "波动率组合"
    if any(keyword in name for keyword in ("价差", "Ratio")):
        return "方向价差"
    return "单腿方向"


def objective_tags(name: str, family: str) -> list[str]:
    tags: list[str] = []
    if family == "套利":
        tags.append("套利")
    if family == "对冲/收益增强":
        if any(keyword in name for keyword in ("保护性", "领口", "双限")):
            tags.append("风险对冲")
        if "备兑" in name:
            tags.append("收益增强")
    if family in {"单腿方向", "方向价差"}:
        tags.append("方向表达")
    if family in {"波动率组合", "时间价差"}:
        tags.append("波动率交易")
    if family == "合成复制":
        tags.append("仓位替代")
    return dedupe(tags or ["结构表达"])


def direction_tags(name: str, scenario: str, use_case: str, family: str) -> list[str]:
    if family == "套利":
        return ["套利"]

    corpus = " ".join([scenario, use_case])
    tags: list[str] = []
    if any(keyword in corpus for keyword in ("看涨", "看多", "牛市", "偏多", "上涨")):
        tags.append("看涨")
    if any(keyword in corpus for keyword in ("看跌", "看空", "熊市", "偏空", "下跌")):
        tags.append("看跌")
    if any(keyword in corpus for keyword in ("震荡", "横盘", "窄幅")):
        tags.append("震荡")
    if any(keyword in corpus for keyword in ("方向不明", "双向", "大幅波动")):
        tags.append("双向波动")

    if not tags and family in {"单腿方向", "方向价差", "波动率组合", "时间价差", "合成复制"}:
        fallback_corpus = name
        if any(keyword in fallback_corpus for keyword in ("看涨", "牛市")):
            tags.append("看涨")
        if any(keyword in fallback_corpus for keyword in ("看跌", "熊市")):
            tags.append("看跌")
        if any(keyword in fallback_corpus for keyword in ("跨式", "宽跨式", "蝶", "鹰", "Strap", "Strip")):
            tags.append("双向波动")

    if family == "对冲/收益增强":
        tags.append("对冲")
    return dedupe(tags or ["中性"])


def volatility_tags(scenario: str, use_case: str, greeks: str, family: str) -> list[str]:
    corpus = " ".join([scenario, use_case, greeks, family])
    tags: list[str] = []
    if any(keyword in corpus for keyword in ("波动率上行", "隐含波动率上升", "Vega>0")):
        tags.append("波动率上行")
    if any(keyword in corpus for keyword in ("波动率下行", "Vega<0")):
        tags.append("波动率下行")
    if any(keyword in corpus for keyword in ("波动率中性", "Vega≈0", "Vega 中性")):
        tags.append("波动率中性")
    if any(keyword in corpus for keyword in ("高波动率", "大幅波动")):
        tags.append("高波动")
    if any(keyword in corpus for keyword in ("套利空间", "定价偏差")) or family == "套利":
        tags.append("定价偏差")
    return dedupe(tags or ["波动率中性"])


def time_value_profile(greeks: str, use_case: str) -> str:
    corpus = " ".join([greeks, use_case])
    if "Theta>0" in corpus or "时间价值衰减收益" in corpus:
        return "卖时间"
    if "Theta<0" in corpus:
        return "买时间"
    return "时间中性"


def holding_requirement(name: str, margin: str, use_case: str, family: str) -> str:
    corpus = " ".join([name, margin, use_case])
    if any(keyword in corpus for keyword in ("持有标的股票", "持仓股票", "长期持有标的", "股票买入价")) or any(
        keyword in name for keyword in ("备兑看涨", "保护性看跌", "领口", "双限")
    ):
        return "需持有现货"
    if any(keyword in corpus for keyword in ("卖空标的股票", "卖空股票", "股票卖空价")) or any(
        keyword in name for keyword in ("备兑看跌", "保护性看涨")
    ):
        return "需卖空现货"
    if family == "套利":
        return "机构/套利账户"
    return "无持仓要求"


def capital_profile(margin: str, family: str, holding: str) -> str:
    if holding in {"需持有现货", "需卖空现货"}:
        return "现货覆盖型"
    if "全额支付权利金" in margin:
        return "权利金型"
    if any(keyword in margin for keyword in ("需缴纳", "最低每手", "追加保证金", "取两者较高值")):
        return "保证金型"
    if "部分冲抵" in margin:
        return "组合保证金型"
    if family == "套利":
        return "低杠杆套利型"
    return "组合权利金型"


def summarize_margin(margin: str, holding: str, capital: str, family: str) -> str:
    if holding == "需持有现货":
        return "由现货持仓覆盖义务，核心资金占用在现货仓位。"
    if holding == "需卖空现货":
        return "需要卖空现货配合，对账户权限和保证金更敏感。"
    if "全额支付权利金" in margin:
        return "以支付权利金为主，一般无追加保证金。"
    if "取两者较高值" in margin:
        return "裸卖方保证金较高，且会随标的波动动态变化。"
    if "需缴纳行权价差" in margin:
        return "需要缴纳价差保证金，净权利金可部分抵减。"
    if "部分冲抵" in margin:
        return "需要保证金，但组合腿可提供一定冲抵。"
    if family == "套利":
        return "对成本、流动性和执行质量要求更高。"
    if capital == "组合权利金型":
        return "通常以组合建仓成本为主，额外保证金压力相对有限。"
    return "公开版暂未细化保证金规则，实盘前需按券商规则复核。"


def risk_note(max_loss: str, holding: str, capital: str, time_value: str, family: str) -> str:
    if "无限" in max_loss:
        return "理论最大亏损不封顶，必须先控制仓位、保证金和止损纪律。"
    if family == "套利":
        return "套利并不等于无风险，流动性、滑点和交易成本都可能吞掉利润。"
    if capital == "保证金型":
        return "需要持续盯住保证金占用，极端波动时回撤可能快于预期。"
    if holding in {"需持有现货", "需卖空现货"}:
        return "风险来自期权腿和现货腿的共同暴露，不能只盯权利金本身。"
    if time_value == "买时间":
        return "即使方向判断正确，也可能被时间损耗拖累，需要给行情留出兑现窗口。"
    if time_value == "卖时间":
        return "收益常来自时间衰减，但尾部行情会让小赚大亏的结构迅速暴露。"
    return "先确认最大亏损是否真的可承受，再考虑收益空间和执行复杂度。"


def starter_friendly(level: int, max_loss: str, holding: str, family: str) -> bool:
    if "无限" in max_loss:
        return False
    if level > 3:
        return False
    if holding in {"需卖空现货", "机构/套利账户"}:
        return False
    if family in {"套利", "合成复制"}:
        return False
    return True


def summary_text(
    directions: list[str],
    volatilities: list[str],
    time_value: str,
    objectives: list[str],
    family: str,
) -> str:
    if family == "套利":
        return "更适合关注定价偏差的专业账户，收益主要来自执行效率和成本控制。"
    if "风险对冲" in objectives:
        return "适合已有现货仓位的风险管理场景，用结构化期权换取下行保护或波动缓冲。"
    if "收益增强" in objectives:
        return "适合已有持仓想增厚收益的人群，但要接受上行被限制或现货继续波动。"

    if "双向波动" in directions:
        direction_phrase = "适合预期将出现较大波动但方向不明确的场景"
    elif "看涨" in directions and "看跌" not in directions:
        direction_phrase = "适合偏看涨的市场判断"
    elif "看跌" in directions and "看涨" not in directions:
        direction_phrase = "适合偏看跌的市场判断"
    elif "震荡" in directions:
        direction_phrase = "适合区间震荡或横盘判断"
    else:
        direction_phrase = "适合中性或结构化表达"

    if "定价偏差" in volatilities:
        volatility_phrase = "核心收益更依赖定价偏差而非单纯方向。"
    elif "高波动" in volatilities or "波动率上行" in volatilities:
        volatility_phrase = "更受益于隐含波动率抬升或价格摆动放大。"
    elif "波动率下行" in volatilities:
        volatility_phrase = "更适合卖波动或等待波动率回落。"
    else:
        volatility_phrase = "对波动率的依赖相对中性。"

    time_phrase = {
        "买时间": "建仓后需要行情尽快兑现，时间并不站在你这一边。",
        "卖时间": "时间衰减通常是收益来源，但尾部风险也会更集中。",
        "时间中性": "收益更多来自结构设计，而不是单纯赚取时间衰减。",
    }[time_value]
    return f"{direction_phrase}，{volatility_phrase}{time_phrase}"


def short_breakeven(text: str) -> str:
    value = (text or "").strip()
    if not value:
        return "公开版待补充"
    return value


def read_primary_rows(workbook: Any) -> dict[str, dict[str, Any]]:
    catalog: dict[str, dict[str, Any]] = {}
    for sheet_name in PRIMARY_SHEETS:
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            number, strategy_name, scenario, max_profit, max_loss, greeks, risk_label = row[:7]
            if not strategy_name:
                continue
            chinese_name, english_name = split_strategy_name(str(strategy_name))
            normalized = normalize_strategy_name(chinese_name)
            if normalized not in catalog:
                catalog[normalized] = {
                    "strategy_id": f"strategy-{int(number):03d}" if isinstance(number, (int, float)) else f"strategy-{normalized}",
                    "name": chinese_name,
                    "name_en": english_name,
                    "scenario_raw": str(scenario or ""),
                    "use_case_raw": "",
                    "max_profit": str(max_profit or "公开版待补充"),
                    "max_loss": str(max_loss or "公开版待补充"),
                    "greeks": str(greeks or "公开版待补充"),
                    "risk_level": risk_level_from_label(str(risk_label or "3 级（中风险）")),
                    "risk_label": str(risk_label or "3 级（中风险）"),
                    "breakeven": "",
                    "margin_raw": "",
                    "source_sheets": [sheet_name],
                }
                continue

            existing = catalog[normalized]
            existing["scenario_raw"] = str(scenario or existing["scenario_raw"])
            existing["max_profit"] = str(max_profit or existing["max_profit"])
            existing["max_loss"] = str(max_loss or existing["max_loss"])
            existing["greeks"] = str(greeks or existing["greeks"])
            existing["risk_level"] = risk_level_from_label(str(risk_label or existing["risk_label"]))
            existing["risk_label"] = str(risk_label or existing["risk_label"])
            existing["source_sheets"] = dedupe(existing["source_sheets"] + [sheet_name])
    return catalog


def merge_detail_rows(workbook: Any, catalog: dict[str, dict[str, Any]]) -> None:
    next_index = 1000
    for sheet_name in DETAIL_SHEETS:
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            category, strategy_name, breakeven, margin, max_profit, max_loss, use_case = row[:7]
            if not strategy_name:
                continue
            chinese_name, english_name = split_strategy_name(str(strategy_name))
            normalized = normalize_strategy_name(chinese_name)
            normalized = DETAIL_ALIASES.get(normalized, normalized)
            if normalized not in catalog:
                catalog[normalized] = {
                    "strategy_id": f"strategy-extra-{next_index}",
                    "name": chinese_name,
                    "name_en": english_name,
                    "scenario_raw": str(category or ""),
                    "use_case_raw": str(use_case or ""),
                    "max_profit": str(max_profit or "公开版待补充"),
                    "max_loss": str(max_loss or "公开版待补充"),
                    "greeks": "公开版待补充",
                    "risk_level": 3,
                    "risk_label": "3 级（中风险）",
                    "breakeven": str(breakeven or ""),
                    "margin_raw": str(margin or ""),
                    "source_sheets": [sheet_name],
                }
                next_index += 1
                continue

            item = catalog[normalized]
            item["breakeven"] = str(breakeven or item["breakeven"])
            item["margin_raw"] = str(margin or item["margin_raw"])
            item["use_case_raw"] = str(use_case or item["use_case_raw"])
            if item["max_profit"] == "公开版待补充" and max_profit:
                item["max_profit"] = str(max_profit)
            if item["max_loss"] == "公开版待补充" and max_loss:
                item["max_loss"] = str(max_loss)
            item["source_sheets"] = dedupe(item["source_sheets"] + [sheet_name])


def finalize_catalog(catalog: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    public_records: list[dict[str, Any]] = []
    for item in catalog.values():
        family = strategy_family(item["name"])
        objectives = objective_tags(item["name"], family)
        directions = direction_tags(item["name"], item["scenario_raw"], item["use_case_raw"], family)
        vol_tags = volatility_tags(item["scenario_raw"], item["use_case_raw"], item["greeks"], family)
        time_value = time_value_profile(item["greeks"], item["use_case_raw"])
        holding = holding_requirement(item["name"], item["margin_raw"], item["use_case_raw"], family)
        capital = capital_profile(item["margin_raw"], family, holding)
        level = int(item["risk_level"])
        public_records.append(
            {
                "strategy_id": item["strategy_id"],
                "name": item["name"],
                "name_en": item["name_en"],
                "strategy_family": family,
                "objective_tags": objectives,
                "direction_tags": directions,
                "volatility_tags": vol_tags,
                "time_value_profile": time_value,
                "holding_requirement": holding,
                "capital_profile": capital,
                "risk_level": level,
                "risk_label": item["risk_label"],
                "risk_bucket": risk_bucket(level),
                "max_profit": item["max_profit"],
                "max_loss": item["max_loss"],
                "breakeven": short_breakeven(item["breakeven"]),
                "margin_summary": summarize_margin(item["margin_raw"], holding, capital, family),
                "greeks": item["greeks"],
                "summary": summary_text(directions, vol_tags, time_value, objectives, family),
                "risk_note": risk_note(item["max_loss"], holding, capital, time_value, family),
                "starter_friendly": starter_friendly(level, item["max_loss"], holding, family),
                "source_sheets": item["source_sheets"],
            }
        )
    public_records.sort(key=lambda row: (row["risk_level"], row["strategy_family"], row["name"]))
    return public_records


def build_catalog(source_path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(source_path, data_only=True)
    catalog = read_primary_rows(workbook)
    merge_detail_rows(workbook, catalog)
    return finalize_catalog(catalog)


def write_catalog(records: list[dict[str, Any]], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        json.dumps(records, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
